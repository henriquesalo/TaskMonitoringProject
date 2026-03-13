from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages # importando messages para poder mostrar mensagens de erro no login/registro.
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required # decorators = funcao que envolve outra. Nesse caso se nao tiver logado bloqueia acesso.
from tasks.models import Task, Deliverable, PlannedActivity # importando os modelos
from tasks.forms import TaskForm, RegistrationForm # importando o form da task e o de registro.
from django.utils import timezone # importando timezone para poder usar no createTask e respeita o timezone do projeto.
from django.http import JsonResponse, HttpResponse # importando JsonResponse para poder retornar um json no updateTaskStatus e HttpResponse para poder enviar arquivos ao navegador.
from datetime import datetime, timedelta
from openpyxl import Workbook # importando o Workbook que representa um arquivo excel.
import json 

def build_dashboard_data(tasks):
    tracked_tasks = list(tasks.exclude(start_time__isnull=True).exclude(end_time__isnull=True))
    done_tasks = list(tasks.filter(status='done', end_time__isnull=False))

    today = timezone.localdate()

    daily_dates = [today - timedelta(days=i) for i in range(13, -1, -1)]
    daily_hours = {d: 0 for d in daily_dates}
    daily_done = {d: 0 for d in daily_dates}

    # total of all completed tasks (all-time)
    total_done_all = len(done_tasks)

    # additional metrics
    launches = len(tracked_tasks)
    distinct_activities = len({t.title for t in tracked_tasks})

    for task in tracked_tasks:
        end_date = timezone.localdate(task.end_time)
        if end_date in daily_hours:
            daily_hours[end_date] += task.calculate_worked_hours()

    for task in done_tasks:
        end_date = timezone.localdate(task.end_time)
        if end_date in daily_done:
            daily_done[end_date] += 1

    daily_labels = [d.strftime("%d/%m") for d in daily_dates]
    # preserve full precision in values, rounding happens only for presentation
    daily_hours_values = [daily_hours[d] for d in daily_dates]
    daily_done_values = [daily_done[d] for d in daily_dates]

    week_starts = []
    current_week_start = today - timedelta(days=today.weekday())
    for i in range(7, -1, -1):
        week_starts.append(current_week_start - timedelta(weeks=i))

    weekly_hours = {d: 0 for d in week_starts}
    weekly_done = {d: 0 for d in week_starts}

    for task in tracked_tasks:
        end_date = timezone.localdate(task.end_time)
        week_start = end_date - timedelta(days=end_date.weekday())
        if week_start in weekly_hours:
            weekly_hours[week_start] += task.calculate_worked_hours()

    for task in done_tasks:
        end_date = timezone.localdate(task.end_time)
        week_start = end_date - timedelta(days=end_date.weekday())
        if week_start in weekly_done:
            weekly_done[week_start] += 1

    weekly_labels = [f"Sem {d.isocalendar().week:02d}/{d.isocalendar().year}" for d in week_starts]
    weekly_hours_values = [weekly_hours[d] for d in week_starts]
    weekly_done_values = [weekly_done[d] for d in week_starts]

    month_keys = []
    base_month = today.replace(day=1)
    for i in range(11, -1, -1):
        year = base_month.year
        month = base_month.month - i
        while month <= 0:
            month += 12
            year -= 1
        month_keys.append((year, month))

    monthly_hours = {k: 0 for k in month_keys}
    monthly_done = {k: 0 for k in month_keys}

    for task in tracked_tasks:
        end_date = timezone.localdate(task.end_time)
        key = (end_date.year, end_date.month)
        if key in monthly_hours:
            monthly_hours[key] += task.calculate_worked_hours()

    for task in done_tasks:
        end_date = timezone.localdate(task.end_time)
        key = (end_date.year, end_date.month)
        if key in monthly_done:
            monthly_done[key] += 1

    monthly_labels = [f"{m:02d}/{y}" for (y, m) in month_keys]
    monthly_hours_values = [monthly_hours[k] for k in month_keys]
    monthly_done_values = [monthly_done[k] for k in month_keys]

    # Totais exatos baseados no período de 14 dias pra horas,
    # mas número de tarefas concluidas é global (len(done_tasks)).
    total_hours = sum(daily_hours_values)
    total_done = total_done_all
    
    # Média diária considerando apenas dias com horas registradas
    days_with_work = sum(1 for h in daily_hours_values if h > 0)
    avg_daily = total_hours / max(days_with_work, 1)

    # exposable totals
    days_worked = days_with_work
    launches = launches
    distinct_activities = distinct_activities
    
    # Taxa de conclusão: tarefas finalizadas no período dividido pelo total de
    # tarefas que tiveram fim dentro do mesmo intervalo.
    period_start = daily_dates[0]
    period_end = daily_dates[-1]
    tasks_in_period = tasks.filter(end_time__gte=timezone.make_aware(timezone.datetime.combine(period_start, timezone.datetime.min.time())),
                                   end_time__lte=timezone.make_aware(timezone.datetime.combine(period_end, timezone.datetime.max.time())))
    total_tracked_tasks = tasks_in_period.count()
    completion_rate = round((total_done / max(total_tracked_tasks, 1)) * 100, 1) if total_tracked_tasks > 0 else 0

    # helper to format hours
    def fmt(h):
        h_int = int(h)
        mins = int(round((h - h_int) * 60))
        return f"{h_int:02d}:{mins:02d}"

    return {
        'hours': {
            'daily': {'labels': daily_labels, 'values': daily_hours_values},
            'weekly': {'labels': weekly_labels, 'values': weekly_hours_values},
            'monthly': {'labels': monthly_labels, 'values': monthly_hours_values},
        },
        'done': {
            'daily': {'labels': daily_labels, 'values': daily_done_values},
            'weekly': {'labels': weekly_labels, 'values': weekly_done_values},
            'monthly': {'labels': monthly_labels, 'values': monthly_done_values},
        },
        'totals': {
            'hours': total_hours,
            'hours_hhmm': fmt(total_hours),
            'done': total_done,
            'avg_daily': avg_daily,
            'avg_daily_hhmm': fmt(avg_daily),
            'completion_rate': int(completion_rate),
            'launches': launches,
            'days_worked': days_worked,
            'distinct_activities': distinct_activities,
        }
    }

def build_profile_data(tasks):
    """Retorna a distribuição de horas por atividade **do usuário atual** e
    o total por dia da semana (últimos 14 dias).
    """
    today = timezone.localdate()
    week_days = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sab', 'Dom']
    # apenas tarefas que possuem tempos válidos e pertencem ao usuário
    valid = tasks.filter(end_time__isnull=False, start_time__isnull=False)

    activity_hours = {}
    weekday_hours = {d: 0 for d in week_days}

    for t in valid:
        h = t.calculate_worked_hours()
        activity_hours[t.title] = activity_hours.get(t.title, 0) + h

        end_date = timezone.localdate(t.end_time)
        if end_date >= today - timedelta(days=13):
            wd = week_days[end_date.weekday()]
            weekday_hours[wd] += h

    activity_labels = list(activity_hours.keys())
    activity_values = [round(activity_hours[l], 2) for l in activity_labels]

    weekday_labels = week_days
    weekday_values = [round(weekday_hours[d], 2) for d in weekday_labels]

    return {
        'activity': {'labels': activity_labels, 'values': activity_values},
        'weekday': {'labels': weekday_labels, 'values': weekday_values}
    }

def build_team_activities():
    """Agrupa tarefas por título para mostrar tempo total e por perfil.

    Originalmente isto era limitado aos últimos 14 dias, mas os usuários
    esperam que todas as atividades concluídas apareçam (independentemente de
    quando foram feitas). Por isso a lógica de filtragem foi removida.
    """
    # fetch every finished task that has both start/end times
    all_tasks = Task.objects.filter(
        status='done',
        end_time__isnull=False,
        start_time__isnull=False,
    )

    activities = {}
    for task in all_tasks:
        title = task.title
        hours = task.calculate_worked_hours()
        user_name = task.user.get_full_name() or task.user.username

        if title not in activities:
            activities[title] = {
                'total': 0,
                'by_user': {}
            }

        activities[title]['total'] += hours
        if user_name not in activities[title]['by_user']:
            activities[title]['by_user'][user_name] = 0
        activities[title]['by_user'][user_name] += hours

    return activities

def build_user_activities(user):
    """Agrupa tarefas por título para um usuário específico
    
    Mostra apenas as atividades daquele usuário, com informações de tempo total.
    """
    user_tasks = Task.objects.filter(
        user=user,
        status='done',
        end_time__isnull=False,
        start_time__isnull=False,
    )

    activities = {}
    for task in user_tasks:
        title = task.title
        hours = task.calculate_worked_hours()

        if title not in activities:
            activities[title] = {
                'total': 0,
                'by_user': {}
            }

        activities[title]['total'] += hours
        user_name = user.get_full_name() or user.username
        if user_name not in activities[title]['by_user']:
            activities[title]['by_user'][user_name] = 0
        activities[title]['by_user'][user_name] += hours

    return activities

def userLogin(request):
    if request.method == 'POST':
        username = request.POST['username'] # pegando o username do formulario de login
        password = request.POST['password'] # pegando a senha do formulario de login
        user = authenticate(request, username=username, password=password) # autenticando o usuario com o username e senha
        if user is not None: # se o usuario for encontrado
            login(request, user) # loga o usuario no sistema
            return redirect('kanban') # redireciona para a pagina do kanban.
        else:
            messages.error(request, 'Erro ao fazer login. Usuário ou senha inválidos.') # caso haja um erro no login, mostra uma mensagem de erro.
    return render(request, 'users/login.html') # renderizando o template login.html
    
def userLogout(request):
    logout(request) # logout do usuario
    return redirect('/') # redireciona para a pagina de login.

def registro(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid(): # valida senha forte/ username unico e senhas iguais
            user = form.save() # salvando o form que e o proprio usuario no banco
            login(request, user) # faz com que se for valido ele ja logue automaticamente no sistema
            return redirect('kanban') # redireciona para a pagina do kanban.
        else:
            messages.error(request, 'Erro ao registrar usuário. Verifique se as senhas são iguais e se o nome de usuário é único.') # caso haja um erro no registro, mostra uma mensagem de erro.
    else:
        form = RegistrationForm()
    return render(request, 'users/registro.html', {'form': form}) # renderizando o template registro.html e passando o form como contexto.

@login_required
def kanban(request): # criando funcao kanban que envia uma requisicao com o usuario logado.
    tasks = Task.objects.filter(user=request.user) # Filtrando no banco para que cada usuario veja somente suas tasks.
    
    # Filtro para tarefas concluídas por data específica
    filter_date = request.GET.get('filter_date')
    
    done_tasks = tasks.filter(status='done')
    if filter_date and filter_date.strip():  # Verifica se não é vazio ou apenas espaços
        done_tasks = done_tasks.filter(end_time__date=filter_date)
    
    context = { # criando um dicionario HTML para filtrar as tasks por status
        'todo': tasks.filter(status='todo'),
        'doing': tasks.filter(status='doing'),
        'done': done_tasks,
        'blocked': tasks.filter(status='blocked'),
        'filter_date': filter_date if filter_date and filter_date.strip() else '',
    }
    return render(request, 'task/kanban.html', context) # carrega o template, injeta os dados, gera html e envia a resposta http.


@login_required
def admin_dashboard(request):
    """Visão consolidada de todos os usuários disponível apenas para superusers."""
    if not request.user.is_superuser:
        return redirect('admin:login')

    tasks = Task.objects.all()
    dashboard_data = build_dashboard_data(tasks)
    # adiciona métricas extras para Admin
    profiles = Task.objects.values('user').distinct().count()
    total_hours = dashboard_data['totals']['hours']
    avg_per_profile = total_hours / max(profiles, 1)
    dashboard_data['totals']['profiles'] = profiles
    dashboard_data['totals']['avg_per_profile_hhmm'] = f"{int(avg_per_profile)}h {int((avg_per_profile-int(avg_per_profile))*60):02d}m"

    # perfil -> horas por usuário (para gráfico de pizza)
    profile_hours = {}
    valid = tasks.filter(end_time__isnull=False, start_time__isnull=False)
    for t in valid:
        name = t.user.get_full_name() or t.user.username
        profile_hours[name] = profile_hours.get(name, 0) + t.calculate_worked_hours()
    profile_labels = list(profile_hours.keys())
    profile_values = [round(profile_hours[n], 2) for n in profile_labels]

    # reutiliza agregação de atividades / day-of-week global
    profile_data = {
        'profile': {'labels': profile_labels, 'values': profile_values},
        'weekday': build_profile_data(tasks).get('weekday', {})
    }

    team_activities = build_team_activities()

    # Format team activities for JSON
    team_activities_list = [
        {
            'title': title,
            'total': round(data['total'], 2),
            'by_user': {user: round(hours, 2) for user, hours in data['by_user'].items()}
        }
        for title, data in team_activities.items()
    ]

    # Get all deliverables with creator info
    deliverables = [
        {
            'id': d.id,
            'title': d.title,
            'description': d.description or '',
            'hours': d.hours,
            'status': d.status,
            'start_date': str(d.start_date) if d.start_date else None,
            'end_date': str(d.end_date) if d.end_date else None,
            'status_display': d.get_status_display(),
            'hours_hhmm': d.hours_hhmm(),
            'creator_username': d.user.username if d.user else None,
        } for d in Deliverable.objects.all()
    ]

    # Get all planned activities with creator info
    planned = [
        {
            'id': p.id,
            'title': p.title,
            'description': p.description or '',
            'hours': p.hours,
            'priority': p.priority,
            'due_date': str(p.due_date) if p.due_date else None,
            'priority_display': p.get_priority_display(),
            'hours_hhmm': p.hours_hhmm(),
            'creator_username': p.user.username if p.user else None,
        } for p in PlannedActivity.objects.all()
    ]

    return render(request, 'admin_dashboard.html', {
        'dashboard_data': dashboard_data,
        'profile_data': profile_data,
        'team_activities': team_activities_list,
        'deliverables': deliverables,
        'planned': planned,
    })

def createTask(request):
    if request.method == 'POST': # se o metodo for post, cria uma task com os dados do formulario
        form = TaskForm(request.POST) # pegando os dados enviados e colocando na variavel form
        if form.is_valid(): # is.valid() faz o django verificar se os dados sao validos baseado no que foi definido na models.py
            task = form.save(commit=False) # .save() salva direto no banco, mas com o commit falso ele cria sem salvar, pois nesse caso falta atribuir o dono da task.
            task.user = request.user # aqui definimos quem criou a task. Sem isso a task seria criada sem dono.
            task.save() # apos definir o dono, salvamos de fato no banco.
            return redirect('kanban')
        else:
            form = TaskForm() # caso os dados sejam invalidos, cria um novo form vazio.
    return redirect('kanban') # redireciona para a pagina do kanban.

def updateTask(request, task_id): # funcao para conseguir editar uma tarefa
    task = get_object_or_404(Task, id=task_id, user=request.user) # pegando a task pelo id e restringindo para somente ao user que a criou.
    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task) # se o metodo for post nao criar uma nova task mas sim atualizar essa existente
        if form.is_valid():
            form.save()
            return redirect('kanban') # redireciona para a pagina do kanban.
        else:
            form = TaskForm(instance=task) # caso os dados sejam invalidos, o form ja vem preenchido com os dados atuais.
    return redirect('kanban') # redireciona para a pagina do kanban.

def deleteTask(request, task_id):
    task = get_object_or_404(Task, id=task_id, user=request.user)
    if request.method == 'POST':
        task.delete() # deletando a task do banco de dados.
    return redirect('kanban') # redireciona para a pagina do kanban.

def start_task(request, task_id): # a funcao recebe o usuario(request) e a task pelo id dela.
    task = Task.objects.get(id=task_id, user=request.user) # pegando a task pelo id e armazenando na variavel task e restringindo para somente ao user que a criou.
    task.start_time = timezone.now() # definindo o start time como a hora atual.
    task.save()
    return redirect('kanban') # redireciona para a pagina do kanban.

def finish_task(request, task_id):
    task = Task.objects.get(id=task_id, user=request.user) # pegando a task pelo id e armazenando na variavel task e restringindo para somente ao user que a criou.
    task.end_time = timezone.now() # definindo o end time como a hora atual.
    task.status = 'done' # definindo o status como done e movendo automaticamente para sua coluna.
    task.save()
    return redirect('kanban') # redireciona para a pagina do kanban.

@login_required
def updateTaskStatus(request, task_id):
    if request.method == 'POST':
        data = json.loads(request.body) # convertendo o corpo da req para python dict.
        newStatus = data.get('status') # pegando o status enviado.
        task = Task.objects.get(id=task_id, user=request.user) # pegando a task pelo id e restringindo para somente ao user que a criou.
        oldStatus = task.status # guardando o status anterior da task.
        task.status = newStatus # atualizando o status da task.
        if newStatus == 'doing' and not task.start_time:
            task.start_time = timezone.now()
        if newStatus == 'done' and not task.end_time:
            task.end_time = timezone.now()
        # Lógica para quando a task vai para bloqueada
        if newStatus == 'blocked':
            if not task.blocked_at:  # se nao tinha sido bloqueada antes, marca o inicio do bloqueio
                task.blocked_at = timezone.now()
            if task.start_time and not task.end_time:  # se a task tiver sido iniciada mas nao tiver sido finalizada
                task.end_time = timezone.now()  # congela o tempo de fim temporariamente
        # Lógica para quando a task sai de bloqueada
        if oldStatus == 'blocked' and newStatus != 'blocked':
            if task.blocked_at:  # se tiver data de quando foi bloqueada
                blocked_hours = task.calculate_blocked_duration()  # calcula quanto tempo ficou bloqueada
                task.total_blocked_hours = (task.total_blocked_hours or 0) + blocked_hours  # acumula no total de bloqueio
                task.blocked_at = None  # limpa a data de inicio do bloqueio
            if newStatus == 'doing': # Se voltou para "doing", reseta o end_time para continuar contando
                task.end_time = None
        
        task.save() # salvando a task com o novo status no banco.
        return JsonResponse({'success': True}) # retornando um json com sucesso.

@login_required
def dashboard(request):
    tasks = Task.objects.filter(user=request.user)
    dashboard_data = build_dashboard_data(tasks)
    profile_data = build_profile_data(tasks)
    user_activities = build_user_activities(request.user)    
    # Todos os usuários podem ver entregáveis e atividades previstas (globais)
    deliverables = [
        {
            'id': d.id,
            'title': d.title,
            'description': d.description or '',
            'hours': d.hours,
            'status': d.status,
            'start_date': str(d.start_date) if d.start_date else None,
            'end_date': str(d.end_date) if d.end_date else None,
            'status_display': d.get_status_display(),
            'hours_hhmm': d.hours_hhmm(),
            'creator_name': d.user.get_full_name() if d.user else 'Sistema',
        } for d in Deliverable.objects.all()
    ]
    planned = [
        {
            'id': p.id,
            'title': p.title,
            'description': p.description or '',
            'hours': p.hours,
            'priority': p.priority,
            'due_date': str(p.due_date) if p.due_date else None,
            'priority_display': p.get_priority_display(),
            'hours_hhmm': p.hours_hhmm(),
            'creator_name': p.user.get_full_name() if p.user else 'Sistema',
        } for p in PlannedActivity.objects.all()
    ]
    
    # Format user activities for JSON
    user_activities_list = [
        {
            'title': title,
            'total': round(data['total'], 2),
            'by_user': {user: round(hours, 2) for user, hours in data['by_user'].items()}
        }
        for title, data in user_activities.items()
    ]
    
    return render(request, 'task/dashboard.html', {
        'dashboard_data': dashboard_data,
        'profile_data': profile_data,
        'team_activities': user_activities_list,
        'deliverables': deliverables,
        'planned': planned,
    })

@login_required
def exportDashboardExcel(request):
    tasks = Task.objects.filter(user=request.user)
    dashboard_data = build_dashboard_data(tasks)

    workbook = Workbook()
    workbook.remove(workbook.active)

    def add_sheet(title, labels, values):
        ws = workbook.create_sheet(title=title)
        ws.append(["Período", "Valor"])
        for label, value in zip(labels, values):
            ws.append([label, value])

    add_sheet("Horas - Dia", dashboard_data['hours']['daily']['labels'], dashboard_data['hours']['daily']['values'])
    add_sheet("Horas - Semana", dashboard_data['hours']['weekly']['labels'], dashboard_data['hours']['weekly']['values'])
    add_sheet("Horas - Mês", dashboard_data['hours']['monthly']['labels'], dashboard_data['hours']['monthly']['values'])
    add_sheet("Concluídas - Dia", dashboard_data['done']['daily']['labels'], dashboard_data['done']['daily']['values'])
    add_sheet("Concluídas - Semana", dashboard_data['done']['weekly']['labels'], dashboard_data['done']['weekly']['values'])
    add_sheet("Concluídas - Mês", dashboard_data['done']['monthly']['labels'], dashboard_data['done']['monthly']['values'])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="dashboard.xlsx"'
    workbook.save(response)
    return response

@login_required
def exportTasksExcel(request):
    workbook = Workbook() # criando um novo arquivo excel vazio.


# --- CRUD endpoints para administradores ---
from django.views.decorators.http import require_http_methods

@login_required
@require_http_methods(['GET','POST'])
def deliverable_list(request):
    if request.method == 'GET':
        data = list(Deliverable.objects.values())
        return JsonResponse(data, safe=False)
    # POST - qualquer usuário logado pode criar
    payload = json.loads(request.body)
    d = Deliverable.objects.create(
        title=payload.get('title',''),
        description=payload.get('description',''),
        hours=payload.get('hours',0),
        status=payload.get('status','ongoing'),
        start_date=payload.get('start_date') or None,
        end_date=payload.get('end_date') or None,
        user=request.user
    )
    return JsonResponse({'id': d.id})

@login_required
@require_http_methods(['PUT','DELETE'])
def deliverable_detail(request, id):
    try:
        d = Deliverable.objects.get(id=id)
    except Deliverable.DoesNotExist:
        return JsonResponse({'error':'not found'}, status=404)
    if request.method == 'PUT':
        payload = json.loads(request.body)
        for field in ['title','description','hours','status','start_date','end_date']:
            if field in payload:
                setattr(d, field, payload[field])
        d.save()
        return JsonResponse({'success': True})
    else:
        d.delete()
        return JsonResponse({'success': True})

@login_required
@require_http_methods(['GET','POST'])
def planned_list(request):
    if request.method == 'GET':
        data = list(PlannedActivity.objects.values())
        return JsonResponse(data, safe=False)
    # POST - qualquer usuário logado pode criar
    payload = json.loads(request.body)
    p = PlannedActivity.objects.create(
        title=payload.get('title',''),
        description=payload.get('description',''),
        hours=payload.get('hours',0),
        priority=payload.get('priority','medium'),
        due_date=payload.get('due_date') or None,
        user=request.user
    )
    return JsonResponse({'id': p.id})

@login_required
@require_http_methods(['PUT','DELETE'])
def planned_detail(request, id):
    try:
        p = PlannedActivity.objects.get(id=id)
    except PlannedActivity.DoesNotExist:
        return JsonResponse({'error':'not found'}, status=404)
    if request.method == 'PUT':
        payload = json.loads(request.body)
        for field in ['title','description','hours','priority','due_date']:
            if field in payload:
                setattr(p, field, payload[field])
        p.save()
        return JsonResponse({'success': True})
    else:
        p.delete()
        return JsonResponse({'success': True})

    worksheet = workbook.active # pegando a planilha ativa do arquivo excel.
    worksheet.title = 'Taferas do Usuário ' + request.user.username # definindo o nome da planilha.

    worksheet.append([
        "Título", "Descrição", "Status", "Data de Criação", "Data de Conclusão", "Horas Trabalhadas"
    ])

    tasks = Task.objects.filter(user=request.user) # Pegando apenas as tasks do usuario logado e salvando numa variavel.
    for task in tasks: # Para cada task do usuario logado, adiciona uma nova linha na planilha.
        
        hours = task.worked_hours_formated()
        
        worksheet.append([
            task.title,
            task.description,
            task.status,
            task.created_at.strftime("%d/%m/%Y %H:%M") if task.start_time else "", # se tiver start time, formata como dd/mm/yyyy HH:MM, senao, vazio.
            task.end_time.strftime("%d/%m/%Y %H:%M") if task.end_time else "", # se tiver end time, formata como dd/mm/yyyy HH:MM, senao, vazio.
            hours
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="tasks.xlsx"' # definindo o nome do arquivo excel que sera baixado.
    workbook.save(response) # salvando o arquivo excel na resposta http.
    return response # retornando a response http que é o arquivo excel para download.
